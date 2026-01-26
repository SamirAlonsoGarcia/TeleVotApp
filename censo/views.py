from ast import If
from datetime import timedelta
from django.http import HttpResponseRedirect, FileResponse, JsonResponse, HttpResponseForbidden
from django.shortcuts import get_object_or_404, redirect, render
from django.contrib.auth import authenticate, login, logout
from django.contrib import messages
from django.contrib.messages import get_messages
from django.contrib.auth.decorators import login_required, user_passes_test
from django.views import View
from django.conf import settings
from django.core.files.storage import default_storage
from django.core.mail import send_mail, EmailMessage
from django.db import IntegrityError, connection, DatabaseError, transaction
from django.db.models import Count, Q #Funcion Contar SQL. Para consultas complejas
from django.db.models.signals import post_save
from django.dispatch import receiver
from django.urls import reverse
from django.utils import timezone
from cryptography.hazmat.primitives.asymmetric import rsa
from cryptography.hazmat.primitives.asymmetric import padding
from cryptography.hazmat.primitives import hashes, serialization
from cryptography.hazmat.backends import default_backend
import base64
import os
import hashlib
import uuid
import json
import random
import math
import openpyxl
from functools import wraps

from urllib3 import request
from .models import CalendarioVotacion, Noticias, RolUsuario, Usuario, Censo, MesaElectoral, Votacion, Voto, Candidatura, Incidencia, CensoUsuario, CensoVotacion, CandidatosNombrados, InscritosVotacion, IntegrantesCandidatura, IntegrantesMesa, Certificado, ComunicacionDirector
from .forms import JuntaNoticiaForm, LoginForm, MesaNoticiaForm, NuevoUsuarioForm, NuevaCandidaturaForm, NuevaVotacionForm, AdminCrearUsuarioForm, MesaNoticiaForm, NuevaIncidenciaForm, JuntaIncidenciaForm, CalendarioVotacionForm, AsignarDirectorCampañaForm

#getobject_or_404 funcion que maneja la salida de una operacion con o sin parametros y que devuelve una pagina de error si no se puede procesar.

def role_required(*allowed_roles):

    #Verifica que request.session['rol_actual'] esté en allowed_roles.
    #allowed_roles son los códigos de rol (admin, mesa, elector, ...).

    def decorator(view_func):
        @wraps(view_func)
        @login_required(login_url='tablonAnuncios')
        def _wrapped(request, *args, **kwargs):
            rol = request.session.get('rol_actual')
            if not rol:
                messages.error(request, 'Debes seleccionar un rol para continuar.')
                return redirect('seleccionar_rol')
            if allowed_roles and rol not in allowed_roles:
                messages.error(request, 'No tienes permisos para acceder a esta sección.')
                # redirige al inicio del rol actual
                return redirect(get_url_inicio_rol(rol))
            return view_func(request, *args, **kwargs)
        return _wrapped
    return decorator

#Pagina inicial para entrar a la aplicacion: Tablon de anuncios con noticias + login

def tablonAnuncios(request):
    lista_noticias = Noticias.objects.all()
    form = LoginForm()

    return render(request, "Tablon_de_Anuncios.html", {'noticias': lista_noticias,'form': form,})

def login_view(request):
    form = LoginForm(request.POST or None)
    noticias = Noticias.objects.exclude(IdVotacionRelacionada=0)
    censos = Censo.objects.all()
    mesas = MesaElectoral.objects.all()
    votaciones = Votacion.objects.exclude(Estado=False)

    if request.method == 'POST':
        form = LoginForm(request.POST)

        if form.is_valid():
            documento_fiscal = form.cleaned_data['documento_fiscal']
            password = form.cleaned_data['password']

            user = authenticate(request, username=documento_fiscal, password=password)

            if user is not None:
                login(request, user)
                roles = get_roles_usuario(user)

                if not roles:
                    logout(request)
                    request.session.flush()
                    messages.error(request, 'No tienes roles asignados. Contacta con el administrador.')
                    return redirect('login')

                # 1 único rol redirigimos a su inicio directamente
                if len(roles) == 1:
                    rol_unico = roles[0]
                    request.session['rol_actual'] = rol_unico
                    request.session['roles_disponibles'] = roles
                    return redirect(get_url_inicio_rol(rol_unico))
                else:
                    request.session['roles_disponibles'] = roles
                    return redirect('seleccionar_rol')
            else:
                messages.error(request, 'Documento Fiscal y/o contraseña incorrectos.')
        else:
            messages.error(request, 'Por favor, corrige los errores en el formulario.')
    else:
        form = LoginForm()

    return render(request, 'Tablon_de_Anuncios.html', {'form': form,'noticias': noticias,'censos': censos,'mesas': mesas,'votaciones': votaciones})

def get_roles_usuario(user):
    return list(RolUsuario.objects.filter(usuario=user).values_list('rol', flat=True))

def get_url_inicio_rol(rol_str: str) -> str:
    """
    Mapea el código de rol a la urlname del dashboard.
    """
    seleccion = {
        'admin': 'Inicio_admin',
        'elector': 'Inicio_votante',
        'mesa': 'Inicio_mesa',
        'director': 'Inicio_director',
        'junta': 'Inicio_junta',
    }
    return seleccion.get(rol_str, 'Inicio_generico')

#Mostrar el texto largo en lugar del rol corto interno
ROL_LABELS = dict(RolUsuario.ROL_USUARIO)

@login_required(login_url='login')
def seleccionar_rol(request):
    roles = request.session.get('roles_disponibles')
    if not roles:
        roles = get_roles_usuario(request.user)
        request.session['roles_disponibles'] = roles

    if not roles:
        messages.error(request, 'No tienes roles asignados. Contacta con el administrador.')
        return redirect('login')

    if len(roles) == 1:
        rol_unico = roles[0]
        request.session['rol_actual'] = rol_unico
        return redirect(get_url_inicio_rol(rol_unico))

    if request.method == 'POST':
        rol_seleccionado = request.POST.get('rol')
        if not rol_seleccionado:
            messages.error(request, 'Debes seleccionar un rol para continuar.')
        elif rol_seleccionado not in roles:
            messages.error(request, 'Rol no válido.')
        else:
            request.session['rol_actual'] = rol_seleccionado
            return redirect(get_url_inicio_rol(rol_seleccionado))

    roles_para_template = [{'code': r, 'label': ROL_LABELS.get(r, r)} for r in roles]
    return render(request, 'seleccion/SeleccionRol.html', {'roles': roles_para_template})

#Paginas de inicio por roles

@role_required('mesa')
def Inicio_mesa(request):
    #buscamos la mesa electoral asignada al usuario actual
    mesa_rel = IntegrantesMesa.objects.filter(usuario=request.user).select_related('mesa__IdVotacion').first()

    if not mesa_rel:
        messages.error(request, "No estás asignado a ninguna mesa electoral, aunque tengas rol de mesa. Contacta con el administrador.")
        return redirect('seleccionar_rol')

    mesa = mesa_rel.mesa
    votacion = mesa.IdVotacion

    calendario = getattr(votacion, "calendario", None)
    fase = None
    rangos = None
    if calendario:
        rangos = calendario_rangos(calendario)
        fase = fase_actual(calendario)
        cal_info = {
            "cal": calendario,
            "fase": fase,           # string o None
            "rangos": rangos,       # dict {fase: (ini, fin)}
        }

    # buscamos el certificado de la mesa si estuviera ya creado
    certificado_mesa = Certificado.objects.filter(propietario_mesa=mesa).first()

    # indicadores para el dashboard de mesa electoral
    total_votos = Voto.objects.filter(IdVotacion=votacion).count()
    incidencias_votacion = Incidencia.objects.filter(IdVotacion=votacion).count()
    incidencias_pendientes = Incidencia.objects.filter(IdVotacion=votacion,IncidenciaSolucionada=False).count()

    return render(request, 'roles/mesa/Inicio_mesa.html', {'mesa': mesa,'votacion': votacion,'certificado_mesa': certificado_mesa,'total_votos': total_votos,'incidencias_votacion': incidencias_votacion,'incidencias_pendientes': incidencias_pendientes,"calendario": calendario, "cal_info": cal_info,})

@role_required('elector')
def Inicio_votante(request):
    userLogged = request.user
    Datos_usuario = {
        'nombre': userLogged.Nombre,
        'apellidos': userLogged.Apellidos,
        'documento_fiscal': userLogged.DocumentoFiscal,
        'coreo_electronico': userLogged.email,
    }

    incidenciasNoSolucionadas = Incidencia.objects.filter(IdUsuario=userLogged,IncidenciaSolucionada=False)

    censos = Censo.objects.filter(censousuario__usuario=userLogged).distinct()

    inscritos = InscritosVotacion.objects.filter(usuario=userLogged)
    votaciones = Votacion.objects.filter(IdVotacion__in=[i.votacion.IdVotacion for i in inscritos])

    certi_usuario = Certificado.objects.filter(propietario_usuario=userLogged)
    mesas = MesaElectoral.objects.filter(integrantesmesa__usuario=userLogged).distinct()

    return render(request,'roles/elector/Inicio.html',{'usuario': Datos_usuario,'incidenciasNoSolucionadas': incidenciasNoSolucionadas,'censos': censos,'votaciones': votaciones,'certificado': certi_usuario,'mesas': mesas,})


@role_required('director')
def Inicio_director(request):
    hoy = timezone.now().date()
    envios_hoy = ComunicacionDirector.objects.filter(director=request.user, fecha_envio__date=hoy).count()
    cand_nombrados = CandidatosNombrados.objects.filter(director=request.user).select_related('votacion', 'candidatura').first()

    if not cand_nombrados:
        messages.error(request, "No tienes una candidatura/votación asignada como director de campaña.")
        return render(request, 'roles/director/Inicio_director.html', {'envios_hoy': envios_hoy})
    
    votacion = cand_nombrados.votacion
    candidatura = cand_nombrados.candidatura

    calendario = getattr(votacion, "calendario", None)
    fase= fase_actual(calendario) if calendario else None

    if request.method == "POST":
        if not calendario:
            messages.error(request, 'La votación no tiene calendario asignado. No se puede enviar comunicaciones por el momento.')
            return redirect('Inicio_director')

        if fase != 'campaña':
            messages.warning(request, 'Solo se pueden enviar comunicaciones durante la fase de campaña.')
            return redirect('Inicio_director')

        if envios_hoy >= 3:
            messages.warning(request, 'Has alcanzado el límite de 3 comunicaciones diarias.')
            return redirect('Inicio_director')
        mensaje = request.POST.get('mensaje').strip()
        imagen = request.FILES.get('imagen_adjunto')

        if not mensaje:
            messages.error(request, 'El mensaje no puede estar vacío.')
            return redirect('Inicio_director')
        
        if imagen:
            nombre_imagen = imagen.name.lower()
            if not nombre_imagen.endswith('.jpg'):
                messages.error(request, 'El archivo adjunto debe ser una imagen (jpg)')
                return redirect('Inicio_director')
        
        ComunicacionDirector.objects.create(director=request.user,mensaje=mensaje,imagen_adjunto=imagen if imagen else None,)

        # aqui enviamos comunicacion al censo de la votación
        try:
            ok, err =enviar_comunicacion_censo(votacion=votacion, mensaje=mensaje, image_file=imagen)
            if not ok:
                messages.warning(request, err or "No se pudo enviar la comunicación.")
            else:
                messages.success(request, 'Comunicación enviada correctamente (salida por consola).')
        except Exception as e:
            messages.error(request, f"Error enviando comunicación: {e}")
        return redirect('Inicio_director')

    return render(request, 'roles/director/Inicio_director.html', {'envios_hoy': envios_hoy,'votacion': votacion,'candidatura': candidatura,})

@role_required('junta')
def Inicio_junta(request):
    # Resumen sencillo para el dashboard
    total_censos = Censo.objects.count()
    total_votaciones = Votacion.objects.count()
    votaciones_activas_conteo = Votacion.objects.filter(Estado=True).count()
    incidencias_pendientes = Incidencia.objects.filter(IncidenciaSolucionada=False).count()

    # Listado de votaciones activas para seleccionar
    votaciones_activas = Votacion.objects.filter(Estado=True).order_by('IdVotacion')

    return render(request, 'roles/junta/Inicio_junta.html', {'total_censos': total_censos,'total_votaciones': total_votaciones,'votaciones_activas': votaciones_activas_conteo,'incidencias_pendientes': incidencias_pendientes, 'votaciones_activas_lista':votaciones_activas})


def es_admin(user):
    return user.is_staff or user.is_superuser
ROL_LABELS = dict(RolUsuario.ROL_USUARIO)

@role_required('admin')
def Inicio_admin(request):
    """
    Pagina Inicio del rol Administrador:
    - Enlace a Django Admin (/admin/)
    - Enlace al alta de usuarios (Junta/Admin)
    """
    return render(request, 'roles/admin/inicio.html', {})

#ROL ADMINISTRADOR
#Funciones correspondientes al rol de administrador

@role_required('admin')
@transaction.atomic
def admin_crear_usuario(request):
    if request.method == 'POST':
        form = AdminCrearUsuarioForm(request.POST)
        if form.is_valid():
            data = form.cleaned_data
            rol_destino = request.POST.get('rol_destino')  # 'junta' o 'admin'
            if rol_destino not in ('junta', 'admin'):
                messages.error(request, 'Debes seleccionar si el usuario será Junta o Administrador.')
            else:
                dni = data['DocumentoFiscal']

                # Comprobar si ya existe usuario con ese DocumentoFiscal
                if Usuario.objects.filter(DocumentoFiscal=dni).exists():
                    messages.error(request, f'Ya existe un usuario con documento {dni}.')
                else:
                    #Generar IdEncriptado igual que cuando importas desde censo
                    clave = crear_hash_unico(data['Nombre'],data['Apellidos'],dni)

                    # Crear usuario
                    usuario = Usuario(
                        Nombre=data['Nombre'],
                        Apellidos=data['Apellidos'],
                        DocumentoFiscal=dni,
                        email=data['email'],
                        username=dni,
                        IdEncriptado=clave,
                    )
                    usuario.set_password(data['password'])
                    usuario.primera_vez = True

                    if rol_destino == 'admin':
                        # Si queremos crear otro admin, marcar is_staff=True
                        usuario.is_staff = True

                    usuario.save()

                    # Asignar rol en tu tabla RolUsuario
                    RolUsuario.objects.create(usuario=usuario, rol=rol_destino)

                    messages.success(request,f'Usuario {dni} creado correctamente con rol {rol_destino}.')
                    return redirect('Inicio_admin')
    else:
        form = AdminCrearUsuarioForm()
    return render(request, 'roles/admin/crear_usuario.html', {'form': form})

@role_required('admin')
def admin_gestion_usuarios(request):
    #Administracion de usuarios en el rol administrador: solo puede asignar desasignar el rol 'admin' y 'junta'.
    # Roles que el admin puede gestionar explícitamente
    allowed_roles = {'admin', 'junta'}

    # Cargamos todos los usuarios
    usuarios = Usuario.objects.all().order_by('DocumentoFiscal')

    # Preparamos un diccionario: { usuario_id: set([roles...]) }
    roles_por_usuario = {
        u.id: set(RolUsuario.objects.filter(usuario=u).values_list('rol', flat=True))
        for u in usuarios
    }

    if request.method == 'POST':
        usuario_id = request.POST.get('usuario_id')
        usuario = get_object_or_404(Usuario, pk=usuario_id)

        # Roles que vienen marcados en el formulario (filtramos solo admin/junta)
        roles_seleccionados = set(request.POST.getlist('roles')) & allowed_roles

        # Roles actuales de ese usuario, pero solo de los que puede gestionar el admin
        roles_actuales = set(RolUsuario.objects.filter(usuario=usuario, rol__in=allowed_roles).values_list('rol', flat=True))

        # Añadir nuevos roles marcados que antes no tenía
        for rol in roles_seleccionados - roles_actuales:
            RolUsuario.objects.get_or_create(usuario=usuario, rol=rol)

        # Quitar roles que antes tenía y ahora se han desmarcado
        RolUsuario.objects.filter(usuario=usuario,rol__in=(roles_actuales - roles_seleccionados)).delete()

        messages.success(request,f'Roles actualizados para el usuario {usuario.DocumentoFiscal}.')
        return redirect('Admin_gestion_usuarios')

    return render(request, 'roles/admin/Gestion_usuarios.html', {'usuarios': usuarios,'roles_por_usuario': roles_por_usuario,})

#ROL JUNTA ELECTORAL
#Funciones correspondientes al rol de junta electoral

@role_required('junta')
def Junta_censos(request):
    censos = Censo.objects.all()
    return render(request, 'roles/junta/Censos.html', {'censos': censos})

@role_required('junta')
def junta_censo_subir_excel(request):
    """
    Sube Excel de censo y crea registro de Censo con fichero asociado.
    Excel esperado (cabeceras en fila 1): DNI | Nombre | Apellidos | Email
    """
    if request.method == 'POST':
        try:
            excel_file = request.FILES['excel_file']
            nombre_censo = request.POST.get('nombre_censo')
            descripcion = request.POST.get('descripcion')

            # Guardar en almacenamiento y leer
            file_path = default_storage.save(f'temp/{excel_file.name}', excel_file)
            abs_path = default_storage.path(file_path)

            wb = openpyxl.load_workbook(abs_path)
            ws = wb.active
            #empieza en la fila dos
            total_censo = ws.max_row - 1 

            nuevo = Censo.objects.create(
                NombreCenso=nombre_censo,
                Descripcion=descripcion,
                nCensados=total_censo,
            )
            # Guarda el fichero en tu campo FileField
            with open(abs_path, 'rb') as f:
                nuevo.FicheroAsociado.save(excel_file.name, f)
            wb.close()
            default_storage.delete(file_path)

            messages.success(request, f'Censo "{nombre_censo}" creado ({total_censo} personas).')
            return redirect('Junta_censos')
        except Exception as ex:
            messages.error(request, f'Error al subir censo: {ex}')
            return redirect('Junta_censos')

    # GET - formulario de subida dentro del propio Censos.html
    return redirect('Junta_censos')

@transaction.atomic
@role_required('junta')
def junta_censo_inscribir_usuarios(request):
    """
    Lee el Excel del censo seleccionado, crea usuarios NO existentes,
    les asigna rol 'elector' (si no lo tienen) y los asocia al censo.
    """
    if request.method == "POST":
        censo_id = request.POST.get("censo_id")
        if not censo_id:
            messages.error(request, "No has seleccionado ningún censo.")
            return redirect('Junta_censos')

        censo = Censo.objects.get(IdCenso=censo_id)
        ruta = censo.FicheroAsociado.path

        wb = openpyxl.load_workbook(ruta)
        hoja = wb.active

        creados, ya_existian, inscritos = 0, 0, 0
        for fila in hoja.iter_rows(min_row=2, values_only=True):
            if not fila or not fila[0]:
                continue
            dni, nombre, apellidos, email = fila[0], fila[1], fila[2], fila[3]
            clave = crear_hash_unico(nombre, apellidos, dni)

            usuario, creado = Usuario.objects.get_or_create(
                username=dni,
                defaults={
                    'Nombre': nombre,
                    'Apellidos': apellidos,
                    'DocumentoFiscal': dni,
                    'email': email,
                    'IdEncriptado': clave,
                }
            )
            if creado:
                creados += 1
                contraseña_inicial = dni[::-1]  # contraseña provisional = DNI al revés
                usuario.set_password(contraseña_inicial)
                if hasattr(usuario, 'primera_vez'):
                    usuario.primera_vez = True
                usuario.save()
            else:
                ya_existian += 1

            # Asegura rol 'elector'
            RolUsuario.objects.get_or_create(usuario=usuario, rol='elector')

            # Asocia al censo
            _, rel_created = CensoUsuario.objects.get_or_create(usuario=usuario, censo=censo)
            if rel_created:
                inscritos += 1

        wb.close()
        messages.success(request, f'Usuarios: nuevos {creados}, existentes {ya_existian}, inscritos {inscritos}.')
        return redirect('Junta_censos')

    return redirect('Junta_censos')

@role_required('junta')
def junta_vot_asignar_censo(request):
    if request.method == "POST":
        censo_id = request.POST.get("censo_id")
        votacion_id = request.POST.get("votacion_id")
        if not censo_id or not votacion_id:
            messages.error(request, "Selecciona censo y votación.")
            return redirect(f"{reverse('Junta_votaciones_gestionar')}?votacion_id={votacion_id or ''}")

        censo = Censo.objects.get(IdCenso=censo_id)
        votacion = Votacion.objects.get(IdVotacion=votacion_id)

        _, creado = CensoVotacion.objects.get_or_create(censo=censo, votacion=votacion)
        if creado:
            messages.success(request, f'Censo "{censo.NombreCenso}" asociado a la votación.')
        else:
            messages.info(request, f'El censo ya estaba asociado.')
        return redirect(f"{reverse('Junta_votaciones_gestionar')}?votacion_id={votacion_id}")

    return redirect('Inicio_junta')

@transaction.atomic
@role_required('junta')
def junta_vot_crear_candidatura(request):
    if request.method == "POST":
        votacion_id = request.POST.get("votacion_id")
        votacion = Votacion.objects.get(IdVotacion=votacion_id)
        form = NuevaCandidaturaForm(request.POST, votacion=votacion)

        if form.is_valid():
            usuarios = list(form.cleaned_data['usuarios'])

            # No repetir usuarios entre candidaturas de la misma votación
            ya_en_candidaturas = IntegrantesCandidatura.objects.filter(
                candidatura__candidatosnombrados__votacion=votacion,
                usuario__in=usuarios
            ).values_list('usuario_id', flat=True).distinct()
            if ya_en_candidaturas:
                messages.error(
                    request,
                    f"Usuarios ya presentes en candidaturas de esta votación: {', '.join(map(str, ya_en_candidaturas))}."
                )
                return redirect(f"{reverse('Junta_votaciones_gestionar')}?votacion_id={votacion_id}")

            # Crear candidatura + integrantes
            candidatura = form.save(commit=False)
            candidatura.save()
            for u in usuarios:
                IntegrantesCandidatura.objects.create(candidatura=candidatura, usuario=u)

            try:
                obj = CandidatosNombrados.objects.create(votacion=votacion, candidatura=candidatura)
                print("CREADO CandidatosNombrados:", obj.id)
            except Exception as e:
                print("ERROR creando CandidatosNombrados:", repr(e))
                raise

            # Vincular candidatura a votación (CandidatosNombrados)
            #CandidatosNombrados.objects.create(votacion=votacion, candidatura=candidatura)

            messages.success(request, "Candidatura creada y asociada.")
        return redirect(f"{reverse('Junta_votaciones_gestionar')}?votacion_id={votacion_id}")

    return redirect('Inicio_junta')

@role_required('junta')
def junta_vot_inscribir_usuarios(request):
    if request.method == "POST":
        votacion_id = request.POST.get("votacion_id")
        if not votacion_id:
            messages.error(request, "No se ha especificado la votación.")
            return redirect('Inicio_junta')
        votacion = get_object_or_404(Votacion, IdVotacion=votacion_id)

        try:
            cv = CensoVotacion.objects.get(votacion=votacion)
        except CensoVotacion.DoesNotExist:
            messages.error(request, "Asocia primero un censo a la votación.")
            return redirect(f"{reverse('Junta_votaciones_gestionar')}?votacion_id={votacion_id}")

        usuarios_censo = CensoUsuario.objects.filter(censo=cv.censo).select_related('usuario')
        nuevos_inscritos = 0
        nuevos_con_rol_elector = 0

        with transaction.atomic():
            for cu in usuarios_censo:
                usuario = cu.usuario
                # inscribir primero en votación
                _, created_inscrito = InscritosVotacion.objects.get_or_create(votacion=votacion,usuario=usuario)
                if created_inscrito:
                    nuevos_inscritos += 1
                # asegurar rol 'elector'
                _, created_rol = RolUsuario.objects.get_or_create(usuario=usuario, rol='elector')
                if created_rol:
                    nuevos_con_rol_elector += 1

        messages.success(request, f"Proceso completado: "
                                    f"{nuevos_inscritos} usuarios inscritos en la votación · "
                                    f"{nuevos_con_rol_elector} nuevos roles 'elector' asignados.")
        
        return redirect(f"{reverse('Junta_votaciones_gestionar')}?votacion_id={votacion_id}")
    else:
        return redirect('Inicio_junta')

@role_required('junta')
def junta_vot_notificar_usuarios(request):
    if request.method == "POST":
        votacion_id = request.POST.get("votacion_id")
        if not votacion_id:
            messages.error(request, "No se ha especificado la votación.")
            return redirect('Inicio_junta')
        votacion = Votacion.objects.get(IdVotacion=votacion_id)

        try:
            cv = CensoVotacion.objects.get(votacion=votacion)
        except CensoVotacion.DoesNotExist:
            messages.error(request, "Asocia primero un censo a la votación.")
            return redirect(f"{reverse('Junta_votaciones_gestionar')}?votacion_id={votacion_id}")

        usuarios_censo = CensoUsuario.objects.filter(censo=cv.censo).select_related('usuario')
        errores = []
        for cu in usuarios_censo:
            ok = enviarNotificacionUsuario(cu.usuario, votacion)
            if not ok:
                errores.append(cu.usuario.username)
        if errores:
            messages.warning(request, f"No se pudo notificar a: {', '.join(errores)}")
        else:
            messages.success(request, "Usuarios notificados correctamente.")

        return redirect(f"{reverse('Junta_votaciones_gestionar')}?votacion_id={votacion_id}")
    else:
        return redirect('Inicio_junta')

@role_required('junta')
def junta_vot_sortear_mesa(request):
    votacion_id = request.POST.get("votacion_id")
    votacion = Votacion.objects.get(IdVotacion=votacion_id)

    mesa_ya = MesaElectoral.objects.filter(IdVotacion=votacion)
    if mesa_ya.exists():
        messages.error(request, f'La/s mesa/s de "{votacion.TituloVotacion}" ya están creadas.')
        return redirect(f"{reverse('Junta_votaciones_gestionar')}?votacion_id={votacion_id}")

    if request.method == "POST":
        cv = CensoVotacion.objects.filter(votacion=votacion).first()
        if not cv:
            messages.error(request, "Asocia primero censo y crea candidaturas antes del sorteo.")
            return redirect(f"{reverse('Junta_votaciones_gestionar')}?votacion_id={votacion_id}")

        # Excluir usuarios en candidaturas
        cand_ids = CandidatosNombrados.objects.filter(votacion=votacion).values_list('candidatura', flat=True)
        if not cand_ids:
            messages.error(request, "Crea candidaturas antes del sorteo.")
            return redirect(f"{reverse('Junta_votaciones_gestionar')}?votacion_id={votacion_id}")

        usuarios_candidatos_ids = set(
            IntegrantesCandidatura.objects.filter(candidatura__in=cand_ids).values_list('usuario', flat=True)
        )
        usuarios_censo_ids = CensoUsuario.objects.filter(censo=cv.censo).values_list('usuario', flat=True)
        elegibles = Usuario.objects.filter(id__in=usuarios_censo_ids).exclude(id__in=usuarios_candidatos_ids)

        total = elegibles.count()
        n_componentes = max(1, math.ceil(total / 25))  # al menos 1
        seleccionados = random.sample(list(elegibles), min(n_componentes, total))

        mesa = MesaElectoral.objects.create(
            IdVotacion=votacion,
            NombreMesa=f"{votacion.TituloVotacion} · Mesa Electoral 1",
            Sorteada=True
        )
        for u in seleccionados:
            IntegrantesMesa.objects.create(usuario=u, mesa=mesa)

        messages.success(request, "Mesa sorteada correctamente.")
        return redirect(f"{reverse('Junta_votaciones_gestionar')}?votacion_id={votacion_id}")

    return redirect('Inicio_junta')

@role_required('junta')
def Junta_vot_asignar_director(request):
    if request.method != "POST":
        return redirect('Junta_votaciones_listado')
        
    votacion_id = request.POST.get("votacion_id")
    if not votacion_id:
        messages.error(request, "Falta id_votacion.")
        return redirect('Junta_votaciones_listado')
    votacion = get_object_or_404(Votacion, IdVotacion=votacion_id)
    candidaturas_id = CandidatosNombrados.objects.filter(votacion=votacion).values_list('candidatura', flat=True).distinct()
    candidaturas_obj = Candidatura.objects.filter(IdCandidatura__in=candidaturas_id).order_by('NombreCandidatura')

    form=AsignarDirectorCampañaForm(request.POST, candidaturas_qs=candidaturas_obj)
    if not form.is_valid():
        messages.error(request, "Revisa los datos del formulario (candidatura/usuario).")
        return redirect(f"{reverse('Junta_votaciones_gestionar')}?votacion_id={votacion_id}")
    
    candidatura = form.cleaned_data['candidatura']
    director = form.cleaned_data['director']

    candidato_nombrado = get_object_or_404(CandidatosNombrados, votacion=votacion, candidatura=candidatura)
    candidato_nombrado.director = director
    ya_es_director = CandidatosNombrados.objects.filter(votacion=votacion,director=director).exclude(pk=candidato_nombrado.pk).exists()

    if ya_es_director:
        messages.error(
            request,
            f"El usuario {director.DocumentoFiscal} ya es director de otra candidatura en esta votación."
        )
        return redirect(f"{reverse('Junta_votaciones_gestionar')}?votacion_id={votacion_id}")

    candidato_nombrado.save()

    RolUsuario.objects.get_or_create(usuario=director, rol='director')
    messages.success(request, f"Director asignado: {director.DocumentoFiscal} → {candidatura.NombreCandidatura}")

    return redirect(f"{reverse('Junta_votaciones_gestionar')}?votacion_id={votacion_id}")

@role_required('junta')
def Censos_junta(request):
    censos = Censo.objects.all().order_by('IdCenso')  # ordenados por id del censo mostraremos tmb el titulo.
    return render(request, 'roles/junta/Censos.html', {'censos': censos,})

@role_required('junta')
def inscribir_usuarios_censo(request):
    if request.method == "POST":
        censo_id = request.POST.get("censo_id")

        if not censo_id:
            messages.error(request, "No has seleccionado ningún censo.")
            return redirect('Junta_censos')

        try:
            censo = Censo.objects.get(IdCenso=censo_id)
        except Censo.DoesNotExist:
            messages.error(request, "El censo seleccionado no existe.")
            return redirect('Junta_censos')

        fichero_ruta = censo.FicheroAsociado.path
        wb = openpyxl.load_workbook(fichero_ruta)
        hoja = wb.active

        # Guardamos las metricas de usuarios para saber cuantos se han creado, ya existian e inscritos
        creados = 0
        ya_existian = 0
        inscritos = 0

        for fila in hoja.iter_rows(min_row=2, values_only=True):
            if not fila or not fila[0]:
                continue

            Dni = fila[0]
            nombre = fila[1]
            apellidos = fila[2]
            email = fila[3]
            clave_encriptada = crear_hash_unico(nombre, apellidos, Dni)

            usuario, creado = Usuario.objects.get_or_create(
                username=Dni,
                defaults={
                    'Nombre': nombre,
                    'Apellidos': apellidos,
                    'DocumentoFiscal': Dni,
                    'email': email,
                    'IdEncriptado': clave_encriptada,
                }
            )

            if creado:
                creados += 1
            else:
                ya_existian += 1

            # Registrar al usuario creado o seleccionado como elector
            RolUsuario.objects.get_or_create(
                usuario=usuario,
                rol='elector'
            )

            # Relación usuario–censo
            _, relacion_creada = CensoUsuario.objects.get_or_create(
                usuario=usuario,
                censo=censo
            )
            if relacion_creada:
                inscritos += 1

        wb.close()
        #informamos de los resultados del procesamiento del fichero de censo
        messages.success(
            request,
            f'Proceso completado. Usuarios nuevos: {creados}, ya existentes: {ya_existian}, '
            f'inscritos en el censo: {inscritos}.'
        )
        return redirect('Junta_censos')

    # Si entra por GET u otro método redireccionamos a la pagina inicial de censos
    return redirect('Junta_censos')

@role_required('junta')
def junta_votaciones_gestionar(request):

    votacion_id = request.POST.get('votacion_id') or request.GET.get('votacion_id')
    if not votacion_id:
        messages.error(request, "No has seleccionado ninguna votación.")
        return redirect('Junta_votaciones_listado')

    votacion = get_object_or_404(Votacion, IdVotacion=votacion_id)

    calendario = CalendarioVotacion.objects.filter(votacion=votacion).first()

    calendario_bloqueado = False
    if calendario and timezone.now().date() > calendario.fecha_no_modificacion:
        calendario_bloqueado = True

    if request.method == "POST" and request.POST.get("accion") == "guardar_calendario":

        if calendario_bloqueado:
            messages.error(request, "El calendario ya está bloqueado y no puede modificarse.")
            return redirect(f"{reverse('Junta_votaciones_gestionar')}?votacion_id={votacion_id}")

        form_calendario = CalendarioVotacionForm(request.POST, instance=calendario)

        if form_calendario.is_valid():
            cal = form_calendario.save(commit=False)
            cal.votacion = votacion

            # Fecha límite de modificación: por defecto mañana
            if not cal.fecha_no_modificacion:
                cal.fecha_no_modificacion = timezone.now().date() + timezone.timedelta(days=1)

            cal.save()
            messages.success(request, "Calendario de la votación guardado correctamente.")
            return redirect(f"{reverse('Junta_votaciones_gestionar')}?votacion_id={votacion_id}")
        else:
            messages.error(request, "Hay errores en el calendario. Revisa los datos.")
    else:
        form_calendario = CalendarioVotacionForm(instance=calendario)

    censos = Censo.objects.all()
    form_candidatura = NuevaCandidaturaForm(votacion=votacion)
    mesa_ya_sorteada = MesaElectoral.objects.filter(IdVotacion=votacion).exists()

    candidaturas_votacion = (CandidatosNombrados.objects.filter(votacion=votacion).select_related('candidatura').values_list('candidatura', flat=True).distinct())

    candidaturas = Candidatura.objects.filter(IdCandidatura__in=candidaturas_votacion)

    integrantes_asociados = {}
    for cand in candidaturas:
        integrantes = (IntegrantesCandidatura.objects.filter(candidatura=cand).select_related('usuario'))
        integrantes_asociados[cand.IdCandidatura] = [
            i.usuario.DocumentoFiscal for i in integrantes
        ]
    
    form_director = AsignarDirectorCampañaForm(candidaturas_qs=candidaturas)

    return render(request,'roles/junta/ManejarVotacion.html',{'votacion': votacion,'censos': censos,'form1': form_candidatura,'mesa_ya_sorteada': mesa_ya_sorteada,'candidaturas': candidaturas,'integrantes': integrantes_asociados,'calendario': calendario,'form_calendario': form_calendario,'calendario_bloqueado': calendario_bloqueado, 'form_director': form_director,})

@role_required('junta')
def junta_vot_eliminar_candidatura(request):
    #hay que tocar todo esto
    if request.method == "POST":
        votacion_id = request.POST.get("votacion_id")
        candidatura_id = request.POST.get("candidatura_id")
        if not votacion_id or not candidatura_id:
            messages.error(request, "Faltan datos para eliminar la candidatura.")
            return redirect(f"{reverse('Junta_votaciones_gestionar')}?votacion_id={votacion_id or ''}")

        votacion = Votacion.objects.get(IdVotacion=votacion_id)
        candidatura = Candidatura.objects.get(IdCandidatura=candidatura_id)

        # Eliminar CandidatosNombrados
        CandidatosNombrados.objects.filter(votacion=votacion, candidatura=candidatura).delete()
        # Eliminar IntegrantesCandidatura
        IntegrantesCandidatura.objects.filter(candidatura=candidatura).delete()
        # Eliminar Candidatura
        candidatura.delete()

        messages.success(request, "Candidatura eliminada correctamente.")
        return redirect(f"{reverse('Junta_votaciones_gestionar')}?votacion_id={votacion_id}")

    return redirect('Inicio_junta')

@role_required('junta')
def junta_votaciones_listado(request):
    votaciones = Votacion.objects.all().order_by('-IdVotacion')

    # formulario de creación, por defecto vacío
    form_nueva = NuevaVotacionForm()

    if request.method == 'POST':
        accion = request.POST.get('accion')

        # FORMULARIO PARA CREAR NUEVA VOTACIÓN DENTRO DE LA MISMA PAGINA
        if accion == 'crear':
            form_nueva = NuevaVotacionForm(request.POST, request.FILES)
            if form_nueva.is_valid():
                nueva = Votacion.objects.create(
                    TituloVotacion=form_nueva.cleaned_data['tituloVotacion'],
                    NParticipantes=form_nueva.cleaned_data['numeroParticipantes'],
                    BasesVotacion=form_nueva.cleaned_data['basesVotacion'],
                    Descripcion=form_nueva.cleaned_data['descripcion'],
                    # Resultado, Estado, RecuentoAutorizado → defaults del modelo
                )
                messages.success(request, f'Votación "{nueva.TituloVotacion}" creada correctamente.')

                # Opcional: ir directamente a manejar esa votación
                url = reverse('Junta_votaciones_gestionar')
                return redirect(f'{url}?votacion_id={nueva.IdVotacion}')
            else:
                messages.error(request, 'Revisa los datos del formulario de nueva votación.')

    return render(request,'roles/junta/Votaciones.html',{'votaciones': votaciones,'form_nueva': form_nueva,})

@role_required('junta')
def junta_incidencias (request):
    #Listado de incidencias para la Junta Electoral.
    incidencias = Incidencia.objects.select_related('IdUsuario', 'IdCenso', 'IdVotacion').order_by('-IdIncidencia')
    incidencia_actual = None
    form=None
    if request.method == 'POST':
        incidencia_id = request.POST.get('incidencia_id')
        if not incidencia_id:
            messages.error(request, "No has seleccionado ninguna incidencia.")
            return redirect('Junta_incidencias')
        
        incidencia_actual = get_object_or_404(Incidencia, pk=incidencia_id)
        form = NuevaIncidenciaForm(request.POST, instance=incidencia_actual)
        if form.is_valid():
            form.save()
            messages.success(request, "Incidencia actualizada correctamente.")
            return redirect('Junta_incidencias')
        else:
            messages.error(request, "Revisa los datos de la incidencia.")
    else:
        incidencia_id = request.GET.get('incidencia_id')
        if incidencia_id:
            incidencia_actual = get_object_or_404(Incidencia, pk=incidencia_id)
            form = NuevaIncidenciaForm(instance=incidencia_actual)
    if form is None:
        form = NuevaIncidenciaForm()
    
    return render(request, 'roles/junta/Incidencias.html', {'incidencias': incidencias,'form': form,'incidencia_actual': incidencia_actual,})

@role_required('junta')
def junta_noticias (request):
    #Gestionamos(ver, crear, editar y eliminar) las noticias de la aplicación desde la Junta Electoral, y se muestran en la pantalla principal de la aplicación.
    noticias = Noticias.objects.all().order_by('-IdNoticia')
    noticia_actual = None
    if request.method == 'POST':
        accion = request.POST.get('accion')
        noticia_id = request.POST.get('noticia_id')

        if accion in ('crear', 'guardar'):
            if noticia_id:  # editar
                noticia_actual = get_object_or_404(Noticias, pk=noticia_id)
                form = JuntaNoticiaForm(request.POST, instance=noticia_actual)
            else:          # crear
                form = JuntaNoticiaForm(request.POST)

            if form.is_valid():
                form.save()
                if accion == 'crear':
                    messages.success(request, "Noticia creada correctamente.")
                else:
                    messages.success(request, "Noticia actualizada.")

                return redirect('Junta_noticias')

        elif accion == 'eliminar' and noticia_id:
            noticia = get_object_or_404(Noticias, pk=noticia_id)
            noticia.delete()
            messages.success(request, "Noticia eliminada.")
            return redirect('Junta_noticias')

        else:
            form = JuntaNoticiaForm(request.POST)
    else:
        noticia_id = request.GET.get('noticia_id')
        if noticia_id:
            noticia_actual = get_object_or_404(Noticias, pk=noticia_id)
            form = JuntaNoticiaForm(instance=noticia_actual)
        else:
            form = JuntaNoticiaForm()
    return render(request, 'roles/junta/Noticias.html', {'noticias': noticias,'form': form,'noticia_actual': noticia_actual,})

role_required('junta')
def nueva_votacion(request):
    #Crear una nueva votación desde Junta Electoral.
    if request.method == "POST" :
        form = NuevaVotacionForm(request.POST, request.FILES)
        if form.is_valid():
            # Mapeamos campos del form al modelo Votacion
            Votacion.objects.create(
                TituloVotacion=form.cleaned_data['tituloVotacion'],
                NParticipantes=form.cleaned_data['numeroParticipantes'],
                BasesVotacion=form.cleaned_data['basesVotacion'],
                Descripcion=form.cleaned_data['descripcion'],
                # Resultado, Estado y RecuentoAutorizado se quedan con sus defaults
            )
            messages.success(request, 'Votación creada correctamente.')
            return redirect('Junta_votaciones_listado')
        else:
            messages.error(request, "Error al crear la votación. Revisa los datos.")
    else:
        form= NuevaVotacionForm()

    return render(request, 'roles/junta/NuevaVotacion.html', {'form': form})

#ROL MESA ELECTORAL
#Funciones correspondientes al rol de mesa electoral

@role_required('mesa')
def mesa_certificado(request):
    #Obtener mesa y votación del usuario de mesa
    mesa_rel = IntegrantesMesa.objects.filter(usuario=request.user).select_related('mesa__IdVotacion').first()
    #Comprobamos que no tiene mesa asignada, si no redirigir con error
    if not mesa_rel:
        messages.error(request, "No estás asignado a ninguna mesa electoral.")
        return redirect('seleccionar_rol')

    mesa = mesa_rel.mesa
    votacion = mesa.IdVotacion

    #Certificado asociado a esta mesa (si existe)
    cert = Certificado.objects.filter(propietario_mesa=mesa, TipoCertificado='mesa').first()

    #Comprobacion de que la votación no ha empèzado. Al no haber fechas ahora mismo lo comprobamos por el estado de la votación
    votacion_iniciada = bool(votacion.Estado)

    if request.method == 'POST':
        accion = (request.POST.get('accion') or "").strip()

        # Si la votación ya ha empezado, no permitimos generar ni revocar
        if votacion_iniciada and accion in ('generar', 'revocar'):
            messages.error(request, "No se puede generar o revocar el certificado una vez iniciada la votación.")
            return redirect('Mesa_certificado')

        # Generar certificado
        if accion == 'generar':
            if cert:
                if getattr(cert, 'revocado', False):
                    messages.error(request, "El certificado está revocado. Contacta con la Junta para emitir uno nuevo.")
                else:
                    messages.error(request, "Ya existe un certificado válido para esta mesa.")
                    return redirect('Mesa_certificado')
            
            pub, priv = generar_claves_certificado_mesa()

            Certificado.objects.create(TipoCertificado='mesa',clave_publica=pub,clave_privada=priv,propietario_mesa=mesa)
            messages.success(request, "Certificado generado correctamente.")
                
            return redirect('Mesa_certificado')

        # Revocar certificado (si existe)
        elif accion == 'revocar':
            if not cert:
                messages.error(request, "No hay certificado que revocar.")
                return redirect('Mesa_certificado')
            if getattr(cert, 'revocado', False):
                messages.error(request, "El certificado ya está revocado.")
                return redirect('Mesa_certificado')
            
            motivo = (request.POST.get('motivo_revocacion') or "").strip()
            if not motivo:
                messages.error(request, "Debes indicar un motivo para la revocación.")
                return redirect('Mesa_certificado')

            cert.revocado = True
            cert.motivo_revocacion = motivo
            cert.fecha_revocacion = timezone.now()
            cert.save()
            messages.success(request, "Certificado revocado correctamente.")
            return redirect('Mesa_certificado')

        # Marcar como disponible el certificado para votantes
        elif accion == 'toggle_visible':
            mesa.certificado_visible = not bool(getattr(mesa, 'certificado_visible', False))
            mesa.save(update_fields=['certificado_visible'])
            messages.success(request, "Visibilidad del certificado actualizada.")
            return redirect('Mesa_certificado')
        
        elif accion == 'revelar_privada':
            if not cert or getattr(cert, 'revocado', False):
                messages.error(request, "No hay certificado válido para habilitar.")
                return redirect('Mesa_certificado')

            # Flag temporal en sesión Se podra revelar en la pagina de recuento.
            request.session['mesa_privkey_unlock'] = True
            request.session['mesa_privkey_unlock_ts'] = timezone.now().timestamp()
            messages.warning(request, "Visualización excepcional habilitada durante 5 minutos.")
            return redirect('Mesa_certificado')
        else:
            messages.error(request, "Acción no reconocida.")
            return redirect('Mesa_certificado')

    return render(request, 'roles/mesa/Certificado_mesa.html', {'mesa': mesa,'votacion': votacion,'certificado': cert,'votacion_iniciada': votacion_iniciada,})

@role_required('mesa')
def mesa_recuento(request):
    # Pantalla de recuento de votación para usuario de mesa electoral. Una vez se llegue al limite de tiempo
    mesa_rel = IntegrantesMesa.objects.filter(usuario=request.user).select_related('mesa__IdVotacion').first()
    if not mesa_rel:
        messages.error(request, "No estás asignado a ninguna mesa electoral.")
        return redirect('seleccionar_rol')
    mesa = mesa_rel.mesa
    votacion = mesa.IdVotacion

    certificados = Certificado.objects.filter(propietario_mesa=mesa, TipoCertificado='mesa', revocado=False).first()

    if not certificados or getattr(certificados, 'revocado', False):
        messages.error(request, "No hay un certificado válido de mesa para poder realizar el recuento.")
        return redirect('Mesa_certificado')

    votacion_cerrada = (votacion.Estado is False)
    if not votacion_cerrada:
        messages.error(request, "La votación aún no ha finalizado. No se puede realizar el recuento.")
        return redirect('Inicio_mesa')

    clave_desbloqueada = mesa_clave_privada_desbloqueada(request, segundos=300)
    resultados=[]
    
    if request.method == 'POST':
        if not clave_desbloqueada:
            messages.error(request, "Debes habilitar la visualización/uso excepcional de la clave privada (5 min) desde Certificado.")
            return redirect('Mesa_recuento')
        # Realizar recuento
        resultados = recuentoVotacion(mesa, votacion, certificados.clave_privada)
        messages.success(request, "Recuento realizado correctamente.")
        
        request.session['mesa_privkey_unlock'] = False
        request.session.pop('mesa_privkey_unlock_ts', None)

    return render(request, 'roles/mesa/RecuentoVotacion.html', {'mesa': mesa,'votacion': votacion,'certificados': certificados,'clave_desbloqueada':clave_desbloqueada,'resultados': resultados,})

def recuentoVotacion(mesa, votacion, clave_privada):
    # Realiza el recuento de votos para la mesa y votación dada, usando la clave privada proporcionada.
    # Verifica que el usuario es miembro de la mesa para esa votación
    es_miembro_mesa = MesaElectoral.objects.filter(votacion=votacion, usuario=request.user).exists()


    if not es_miembro_mesa:
        return HttpResponseForbidden("No tienes permiso para autorizar el recuento.")

    # Cambiar el estado de la votación
    if votacion.estado == False:
        votacion.RecuentoAutorizado = True
        votacion.save()
        messages.success(request, "Se ha autorizado el recuento.")
        certificado_mesa = Certificado.objects.filter(propietario_mesa=mesa, TipoCertificado='mesa', revocado=False).first()
        #aqui falta funcionalidad
        #obtener todos los votos asociados a la votacion
        votos = Voto.objects.filter(IdVotacion=votacion)
        return render(request, 'RecuentoVotacion.Html', {'certificados':certificado_mesa, 'votacion':votacion})
    else:
        messages.warning(request, "La votación aún no está cerrada o ya se autorizó el recuento.")
        return redirect(request,'votaciones')

@role_required('mesa')
def mesa_noticias(request):
    #Obtener mesa y votación del usuario de mesa
    #Manejamos la noticia desde la mesa electoral:eliminar, editar y crear noticias.
    usuario = request.user
    mesa_rel = IntegrantesMesa.objects.filter(usuario=usuario).select_related('mesa__IdVotacion').first()
    # Si no tiene mesa asignada, redirigir con error
    if not mesa_rel:
        messages.error(request, "No estás asignado a ninguna mesa electoral.")
        return redirect('Inicio_mesa')

    mesa = mesa_rel.mesa
    votacion = mesa.IdVotacion

    # Obtener las Noticias asociadas a esta votación
    noticias = Noticias.objects.filter(IdVotacionRelacionada=votacion.IdVotacion).order_by('-IdNoticia')

    # Noticia que estamos manejando (editar/eliminar)
    noticia_actual = None
    
    # Crear nueva noticia (publicación de recuento, fin de votación, etc.)
    if request.method == 'POST':
        accion = request.POST.get('accion')
        noticia_id = request.POST.get('noticia_id')

        if accion in ('crear', 'guardar'):
            if noticia_id:
                noticia_actual = get_object_or_404(Noticias, pk=noticia_id, IdVotacionRelacionada=votacion.IdVotacion)
                form = MesaNoticiaForm(request.POST, instance=noticia_actual)
            else:
                form = MesaNoticiaForm(request.POST)
        if form.is_valid():
            noticia_obj = form.save(commit=False)
            noticia_obj.IdVotacionRelacionada = votacion.IdVotacion
            noticia_obj.IdCensoRelacionada = 0
            noticia_obj.NoticiaApp = False  
            noticia_obj.save()
            if accion == 'crear':
                messages.success(request, "Noticia publicada correctamente.")
            else:
                messages.success(request, "Noticia actualizada correctamente.")
            return redirect('Mesa_noticias')
        elif accion == 'eliminar' and noticia_id:
            noticia_actual = get_object_or_404(Noticias, pk=noticia_id, IdVotacionRelacionada=votacion.IdVotacion)
            noticia_actual.delete()
            messages.success(request, "Noticia eliminada correctamente.")
            return redirect('Mesa_noticias')
        else:
            form = MesaNoticiaForm(request.POST)
    else:
        noticia_id = request.GET.get('noticia_id')
        if noticia_id:
            noticia_actual = get_object_or_404(Noticias, pk=noticia_id, IdVotacionRelacionada=votacion.IdVotacion)
            form = MesaNoticiaForm(instance=noticia_actual)
        else:
            form = MesaNoticiaForm()
    return render(request, 'roles/mesa/Noticias_mesa.html', {'votacion': votacion,'noticias': noticias,'form': form,'noticia_actual': noticia_actual,})        

@role_required('mesa')
def mesa_incidencias(request):
    # Obtener mesa y votación del usuario de mesa
    mesa_rel = IntegrantesMesa.objects.filter(usuario=request.user).select_related('mesa__IdVotacion').first()

    #comprobar que tiene mesa asignada si no redirigir con error
    if not mesa_rel:
        messages.error(request, "No estás asignado a ninguna mesa electoral.")
        return redirect('seleccionar_rol')

    mesa = mesa_rel.mesa
    votacion = mesa.IdVotacion

    # Incidencias asociadas a esta votación
    incidencias = Incidencia.objects.filter(IdVotacion=votacion).select_related('IdUsuario').order_by('IncidenciaSolucionada', 'IdIncidencia')
    incidencia_actual = None
    form = None

    if request.method == 'POST':
        inc_id = request.POST.get('incidencia_id')

        if not inc_id:
            messages.error(request, "Debes seleccionar una incidencia.")
            return redirect('Mesa_incidencias')

        try:
            incidencia_actual = Incidencia.objects.get(IdIncidencia=inc_id, IdVotacion=votacion)
        except Incidencia.DoesNotExist:
            messages.error(request, "La incidencia seleccionada no existe o no pertenece a esta votación.")
            return redirect('Mesa_incidencias')
        form = JuntaIncidenciaForm(request.POST, instance=incidencia_actual)
        if form.is_valid():
            form.save()
            messages.success(request, "Incidencia actualizada correctamente.")
            return redirect('Mesa_incidencias')
        else:
            messages.error(request, "Revisa los datos de la incidencia.")
    else:
        incidencia_id = request.GET.get('incidencia_id')
        if incidencia_id:
            try:
                incidencia_actual = Incidencia.objects.get(IdIncidencia=incidencia_id, IdVotacion=votacion)
                form = JuntaIncidenciaForm(instance=incidencia_actual)
            except Incidencia.DoesNotExist:
                messages.error(request, "La incidencia seleccionada no existe o no pertenece a esta votación.")
    if form is None:
        form = JuntaIncidenciaForm()

    return render(request, 'roles/mesa/Incidencias_mesa.html', {'mesa': mesa,'votacion': votacion,'incidencias': incidencias,'form': form,'incidencia_actual': incidencia_actual,})

#ROL ELECTOR
#Funciones correspondientes al rol de elector

@role_required('elector')
def Censo_view(request):
    censos = Censo.objects.filter(censousuario__usuario=request.user).distinct()

    return render(request,'roles/elector/Censos.html',{'censos': censos})

@role_required('elector')
def Certificados(request):
    userLogged = request.user

    certificado_usuario = Certificado.objects.filter(propietario_usuario=userLogged)

    votaciones_inscrito_activo = InscritosVotacion.objects.filter(usuario=userLogged,votacion__Estado=True).values_list('votacion', flat=True)

    certificados_mesas = Certificado.objects.filter(propietario_mesa__IdVotacion__in=votaciones_inscrito_activo)

    return render(request,'roles/elector/Certificados.html',{'certificado': certificado_usuario,'certificados_mesas': certificados_mesas,})

@role_required('elector')
def Votacion_view(request):
    #Solo cargamos las votaciones en las que el usuario está inscrito
    inscritas = InscritosVotacion.objects.filter(usuario=request.user).values_list('votacion', flat=True)

    votaciones_activas = Votacion.objects.filter(IdVotacion__in=inscritas,Estado=True)

    votaciones_finalizadas = Votacion.objects.filter(IdVotacion__in=inscritas,Estado=False)

    todas = list(votaciones_activas) + list(votaciones_finalizadas)

    # Se buscan todos los calendarios de votación asociados a las votaciones encontradas
    cal_map = {}
    if todas:
        cal_qs = CalendarioVotacion.objects.filter(votacion__in=todas).select_related('votacion')
        hoy = timezone.now().date()
        for cal in cal_qs:
            cal_map[cal.votacion.IdVotacion] = {
                "cal": cal,
                "fase": fase_actual(cal, hoy),
                "rangos": calendario_rangos(cal),
            }

    return render(request,'roles/elector/Votaciones.html',{'votaciones_activas': votaciones_activas,'votaciones_finalizadas': votaciones_finalizadas,"cal_map": cal_map,})

@role_required('elector')
def MandarIncidenciaVotacion(request):
    if request.method!="POST" :
        return redirect('Elector_votaciones')

    titulo = (request.POST.get('titulo') or "").strip()
    texto = (request.POST.get('texto') or "").strip()
    votacion_id = request.POST.get('votacion_id')

    if not votacion_id:
        messages.error(request, "No has seleccionado ninguna votación.")
        return redirect('Elector_votaciones')

    if not titulo or not texto:
        messages.error(request, "Debes indicar un título y una descripción para la incidencia.")
        return redirect('Elector_votaciones')

    try:
        votacion = Votacion.objects.get(IdVotacion=votacion_id)
    except Votacion.DoesNotExist:
        messages.warning(request, "La votación seleccionada no existe.")
    
    if not InscritosVotacion.objects.filter(votacion=votacion, usuario=request.user).exists():
        messages.error(request, "No estás inscrito en la votación seleccionada.")
        return redirect('Elector_votaciones')

    Incidencia.objects.create(IdUsuario=request.user,IdCenso=None,IdVotacion=votacion,TituloIncidencia=titulo,IncidenciaSolucionada=False,TextoIncidencia=texto,)

    messages.success(request, "Incidencia registrada correctamente. La Junta Electoral revisará tu caso.")

    return redirect('Elector_incidencias')

@role_required('elector')
def emitir_voto(request):
    if request.method == 'POST':
        votacion_id = request.POST.get('votacion_id')
        votacion = Votacion.objects.get(IdVotacion=votacion_id)

        candid_votacion = CandidatosNombrados.objects.filter(votacion=votacion)
        candidaturas = Candidatura.objects.filter(IdCandidatura__in=candid_votacion.values_list('candidatura__IdCandidatura', flat=True))

        return render(request,'roles/elector/EmitirVotacion.html',{'votacion': votacion,'candidaturas': candidaturas,})

    # Si entra por GET sin POST, va a votaciones
    return redirect('Elector_votaciones')

@role_required('elector')
def estadisticasVotacion(request):
    votacion_id = request.POST.get("votacion_id")
    votacion = Votacion.objects.get(IdVotacion=votacion_id)

    total_esperado = votacion.NParticipantes
    total_emitidos = Voto.objects.filter(IdVotacion=votacion).count()
    participacion = (total_emitidos / total_esperado) * 100 if total_esperado else 0
    ya_votado = Voto.objects.filter(IdVotacion=votacion,idUsuario=request.user).exists()

    return render(request,'roles/elector/EstadisticasVotacion.html',{'votacion': votacion,'total_esperado': total_esperado,'total_emitidos': total_emitidos,'participacion': participacion,'ya_votado': ya_votado,})

@role_required('elector')
def Noticia(request):
    noticias = Noticias.objects.all()
    return render(request, 'Noticias.html', {'noticias': noticias})

@role_required('elector')
def MandarIncidenciaCenso(request):
    if request.method=="POST" :
        return redirect('Elector_censos')

    titulo = (request.POST.get('titulo') or "").strip()
    texto = (request.POST.get('texto') or "").strip()
    censo_id = request.POST.get('censo_id')  # puede venir vacío

    if not titulo or not texto:
        messages.error(request, "Debes indicar un título y una descripción para la incidencia.")
        return redirect('Elector_censos')

    censo_rel = None
    if censo_id:
        try:
            censo_rel = Censo.objects.get(IdCenso=censo_id)
        except Censo.DoesNotExist:
            messages.warning(request, "El censo seleccionado no existe. Se registrará la incidencia sin censo asociado.")

    Incidencia.objects.create(IdUsuario=request.user,IdCenso=censo_rel,IdVotacion=None,TituloIncidencia=titulo,IncidenciaSolucionada=False,TextoIncidencia=texto,)

    messages.success(request, "Incidencia registrada correctamente. La Junta Electoral revisará tu caso.")

    return redirect('Elector_incidencias')
    
@role_required('elector')
def crear_certificado_personal(request):
    userLogged = request.user
    existe_certificado = Certificado.objects.filter(propietario_usuario=userLogged)
    if not existe_certificado:
        clave_privada = rsa.generate_private_key(public_exponent=65537,key_size=2048)
        clave_privada_encriptada = clave_privada.private_bytes(encoding=serialization.Encoding.PEM,format=serialization.PrivateFormat.TraditionalOpenSSL,encryption_algorithm=serialization.NoEncryption()).decode('utf-8')

        clave_publica_encriptada = clave_privada.public_key().public_bytes(encoding=serialization.Encoding.PEM,format=serialization.PublicFormat.SubjectPublicKeyInfo).decode('utf-8')

        Certificado.objects.create(TipoCertificado='usuario',clave_publica=clave_publica_encriptada,clave_privada=clave_privada_encriptada,propietario_usuario = userLogged)
        messages.success(request,f"Se ha asignado un certificado propio de la aplicacion al usuario {userLogged.DocumentoFiscal}")
    else:
        messages.error(request,f"El usuario: {userLogged.DocumentoFiscal}, ya tiene un certificado asignado. No es posible volver a crearlo")

    return redirect('Elector_certificados')

role_required('elector')
def certificar_voto(request):
    if request.method == "POST":
        usuario = request.user
        candidatura_id = request.POST.get("candidatura_id")
        candidatura = Candidatura.objects.get(IdCandidatura=candidatura_id)
        candidatura_votacion=CandidatosNombrados.objects.get(candidatura=candidatura)
        votacion = Votacion.objects.get(IdVotacion=candidatura_votacion.votacion.IdVotacion)

        voto_creado=Voto.objects.filter(IdVotacion=votacion, idUsuario=usuario)

        if not voto_creado:
            claves_firma=obtenerCertificados(usuario, votacion.IdVotacion)

            voto_cont = f"{candidatura_votacion.votacion.IdVotacion}|{candidatura_votacion.candidatura.IdCandidatura}"

            voto_cifrado= cifrar_clave_publica(voto_cont, claves_firma["cla_mesa_publica"])
            voto_cif_usua= cifrar_con_clave_privada(voto_cont, claves_firma["cla_usuario_privado"])

            Voto.objects.create(
                IdVotacion=votacion,
                idUsuario=usuario,
                hashVoto=voto_cifrado + " :: " + voto_cif_usua
            )
            messages.success(request,"SU VOTO SE EMITIÓ CORRECTAMENTE")
        else:
            messages.success(request,"YA HAS VOTADO EN LA VOTACION ACTUAL, NO ES POSSIBLE VOLVER A VOTAR")
    return redirect('Elector_votaciones')


role_required("elector")
def Noticia(request):
    usuario = request.user
    # Censos donde está inscrito el usuario
    censos_ids = list(CensoUsuario.objects.filter(usuario=usuario).values_list('censo__IdCenso', flat=True))
    # Votaciones en las que participa
    votaciones_ids = list(InscritosVotacion.objects.filter(usuario=usuario).values_list('votacion__IdVotacion', flat=True))
    # Noticias relevantes para este usuario, usamos Q para hacer OR en los filtros
    noticias = Noticias.objects.filter(Q(NoticiaApp=True) | Q(IdCensoRelacionada__in=censos_ids) | Q(IdVotacionRelacionada__in=votaciones_ids)).order_by('-IdNoticia')

    return render(request,'roles/elector/Noticias.html',{'noticias': noticias})

@role_required('elector')
def elector_incidencias(request):
    #Listado de incidencias creadas por el elector.
    incidencias = (Incidencia.objects.filter(IdUsuario=request.user).select_related('IdCenso', 'IdVotacion').order_by('-IdIncidencia'))

    return render(request, 'roles/elector/Incidencias.html', {'incidencias': incidencias,})

#FUNCIONES QUE NO NECESITAN DE DECORADOR DE ROL
#Funciones generales que no necesitan de un rol específico

def descargar_plantilla(request):
    file_path = os.path.join(settings.MEDIA_ROOT, 'Plantilla_Censos.xlsx')
    return FileResponse(open(file_path, 'rb'), as_attachment=True, filename='Plantilla_Censos.xlsx')

def crear_hash_unico(nombre,apellidos,dni):
    aleatorio = f"{nombre}{apellidos}{dni}{uuid.uuid4()}"
    return hashlib.sha256(aleatorio.encode()).hexdigest()

def enviarNotificacionUsuario(usuario, votacion):
    try:
        email_usuario = usuario.email
        asunto = f"Inscripción a la votación: {votacion.TituloVotacion}"
        mensaje = (
            f"Hola {usuario.Nombre}, {usuario.Apellidos}\n\n"
            f"Te informamos que has sido inscrito/a correctamente en la votación:\n"
            f"Título: {votacion.TituloVotacion}\n"
            f"Descripción: {votacion.Descripcion}\n\n"
            f"Gracias por participar.\n\n"
            f"Este es un mensaje automático, por favor no respondas."
        )
        send_mail(subject=asunto,message=mensaje,from_email=settings.DEFAULT_FROM_EMAIL,recipient_list=[email_usuario],fail_silently=False)
        return True
    except Exception as e:
        print(f"Error al enviar correo: ({getattr(usuario, 'DocumentoFiscal', usuario)}): {e}")
        return False

def enviar_comunicacion_censo(votacion, mensaje,image_file=None):
    inscritos=(InscritosVotacion.objects.filter(votacion=votacion).select_related('usuario'))
    destinatarios = [i.usuario.email for i in inscritos if i.usuario.email]
    if not destinatarios:
        return False, "No hay destinatarios con correo electrónico."
    
    asunto = f"Comunicación sobre la campaña: {votacion.TituloVotacion}"
    cuerpo_mensaje = (
        f"Se ha publicado una nueva comunicacion sobre la campaña:\n\n"
        f"Se ha publicado una comunicación de campaña para la votación:\n\n"
        f"Título: {votacion.TituloVotacion}\n"
        f"Descripción: {votacion.Descripcion}\n\n"
        f"Mensaje:\n{mensaje}\n\n"
        f"Este es un mensaje automático, por favor no respondas.")
    email = EmailMessage(subject=asunto,body=cuerpo_mensaje,from_email=settings.DEFAULT_FROM_EMAIL,to=destinatarios)
    if image_file:
        email.attach(image_file.name, image_file.read(), image_file.content_type)

    email.send(fail_silently=False)

    return True, None

def generar_claves_certificado_mesa():
    # Genera un par de claves RSA para el certificado de mesa electoral.
    clave_privada = rsa.generate_private_key(public_exponent=65537,key_size=2048)

    privado_pem = clave_privada.private_bytes(encoding=serialization.Encoding.PEM,format=serialization.PrivateFormat.TraditionalOpenSSL,encryption_algorithm=serialization.NoEncryption()).decode('utf-8')

    publico_pem = clave_privada.public_key().public_bytes(encoding=serialization.Encoding.PEM,format=serialization.PublicFormat.SubjectPublicKeyInfo).decode('utf-8')

    return publico_pem, privado_pem

def mesa_clave_privada_desbloqueada(request, segundos=300):
    if not request.session.get('mesa_privkey_unlock'):
        return False
    ts = request.session.get('mesa_privkey_unlock_ts')
    if not ts:
        return False
    try:
        return (timezone.now().timestamp() - float(ts)) <= segundos
    except Exception:
        return False
    
# En guia de usuario como es comun para todos los roles no se pone el decorador de rol_required
def GuiaUsuario(request):
    return render(request, 'Guia_de_Usuario.html')

def cifrar_clave_publica(contenido, clave_publica):
    clave_pub = serialization.load_pem_public_key(clave_publica.encode(),backend=default_backend())
    bytes_mensaje = contenido.encode()
    contenido_cifrado= clave_pub.encrypt(
        bytes_mensaje,
        padding.OAEP(
            mgf=padding.MGF1(algorithm=hashes.SHA256()),
            algorithm=hashes.SHA256(),
            label=None
        )
    )
    return base64.b64encode(contenido_cifrado).decode()

def descifrar_clave_publica(contenido, firma ,clave_publica):
    clave_pub = serialization.load_pem_public_key(
        clave_publica.encode(),
        backend=default_backend()
    )
    val_firma = base64.b64decode(firma.encode())

    try:
        clave_pub.verify(
            val_firma,
            contenido.encode(),
            padding.PKCS1v15(),
            hashes.SHA256()
        )
    except:
        return 0

def cifrar_con_clave_privada(contenido, clave_privada):
    clave_priv = serialization.load_pem_private_key(
        clave_privada.encode(),
        password=None,
        backend=default_backend()
    )
    mensaje_firmado=clave_priv.sign(
        contenido.encode(),
        padding.PKCS1v15(),
        hashes.SHA256()
    )

    return base64.b64encode(mensaje_firmado).decode()

def descifrar_con_clave_privada(contenido_cifrado_base64, clave_privada):
    clave_priv = serialization.load_pem_private_key(
        clave_privada.encode(),
        password=None,
        backend=default_backend()
    )

    contenido_cifrado = base64.b64decode(contenido_cifrado_base64.encode())

    contenido_descifrado = clave_priv.decrypt(
        contenido_cifrado,
        padding.OAEP(
            mgf=padding.MGF1(algorithm=hashes.SHA256),
            algorithm=hashes.SHA256,
            label=None
        )
    )

    return contenido_descifrado.decode()

def cambiar_contraseña(request):
    return render(request, 'CambiarContraseña.html')

def modificar_contraseña(request):
    if request.method !="POST":
        return redirect('login')

    usuario_id = (request.POST.get('usuario_id') or "").strip()
    antigua_contraseña = (request.POST.get('antigua_contraseña') or "").strip()
    nueva_contraseña = (request.POST.get('nueva_contraseña') or "").strip()
    confir_contraseña = (request.POST.get('confirmar_contraseña') or "").strip()
    
    #validar usuario
    try:
        user = Usuario.objects.get(DocumentoFiscal=usuario_id)
    except Usuario.DoesNotExist:
        messages.error(request,"El usuario introducido no existe. Por favor revisa que el usuario sea correcto")
        return redirect('cambiar_contraseña')
    #validar antigua contraseña
    if not user.check_password(antigua_contraseña):
        messages.error(request,"La contraseña anterior no es la correcta, por favor prueba otra vez")
        return redirect('cambiar_contraseña')
    #validar nueva contraseña
    if nueva_contraseña != confir_contraseña :
        messages.error(request,"No has introducido la misma nueva contraseña. Vuelve a intentarlo y asegurate que sean la misma.")
        return redirect('cambiar_contraseña')
    if len(nueva_contraseña) < 6:  
        messages.error(request,"La nueva contraseña debe tener al menos 6 caracteres.")
        return redirect('cambiar_contraseña')
    #actualizamos
    user.set_password(nueva_contraseña)
    user.primera_vez=False
    user.save()
    messages.success(request,f"Se modificó correctamente la contraseña del usuario: {user.DocumentoFiscal}")
    return redirect('login')

def volverInicio(request):
    return redirect('login')

def obtenerCertificados(usuario, votacion_id):
    cert_usuario = Certificado.objects.get(propietario_usuario=usuario)
    mesaVot = MesaElectoral.objects.get(IdVotacion=votacion_id)
    cert_mesa = Certificado.objects.get(propietario_mesa=mesaVot)
    return({"cla_usuario_privado":cert_usuario.clave_privada, "cla_mesa_publica":cert_mesa.clave_publica})

def logout(request):
    logout(request)
    return redirect('login')

def calendario_rangos(cal):
    return {
        "censos": (cal.fecha_censos, cal.fecha_censos + timedelta(days=cal.duracion_censos_dias - 1)),
        "campaña": (cal.fecha_campaña, cal.fecha_campaña + timedelta(days=cal.duracion_campaña_dias - 1)),
        "votacion": (cal.fecha_votacion, cal.fecha_votacion + timedelta(days=cal.duracion_votacion_dias - 1)),
        "recuento": (cal.fecha_recuento, cal.fecha_recuento + timedelta(days=cal.duracion_recuento_dias - 1)),
        "resultados": (cal.fecha_publicacion_resultados, cal.fecha_publicacion_resultados),
    }

def fase_actual(cal, hoy=None):
    if not cal:
        return None
    hoy = hoy or timezone.now().date()
    rangos = calendario_rangos(cal)
    for fase, (ini, fin) in rangos.items():
        if ini <= hoy <= fin:
            return fase
    return None

def rango_fase(inicio, duracion_dias):
    # duracion 1 => inicio == fin
    fin = inicio + timedelta(days=max(duracion_dias, 1) - 1)
    return inicio, fin

    ######## ANTIGUO ##############
    
@login_required
@user_passes_test(es_admin)
def nuevo_usuario(request):
    if request.method == "POST" :
        form = NuevoUsuarioForm(request.POST)
        if form.is_valid():
            nombre_usuario = form.cleaned_data["Nombre"]
            apellidos_usuario = form.cleaned_data["Apellidos"]
            dni_usuario = form.cleaned_data["DocumentoFiscal"]
            email_usuario = form.cleaned_data["Email"]
            clave_encriptada = crear_hash_unico(nombre_usuario,apellidos_usuario,dni_usuario)
        try:
            FunUsuarioNuevo(nombre_usuario,apellidos_usuario,dni_usuario,email_usuario,clave_encriptada)
            messages.success(request, "Usuario Creado Correctamente")
        except IntegrityError as e:
            messages.error(request, "No se ha podido Insertar el usuario, ya esta creado en la BD")
        except Exception as ex:
            messages.error(request, f"No se ha podido Insertar el usuario, revisa los datos {str(ex)}")
            return redirect('administracion')
    else:
        form=NuevoUsuarioForm()
        
    return render(request, 'Administracion.html', {'form': form})

def FunUsuarioNuevo(nombreUsuario, apellidosUsuario, Dni, email, claveUser):
    usuario_nuevo = Usuario.objects.create(
        Nombre=nombreUsuario,
        Apellidos=apellidosUsuario,
        DocumentoFiscal = Dni,
        email=email,
        username=Dni,
        IdEncriptado=claveUser
    )
    usuario_nuevo.save()
    return usuario_nuevo

@login_required
@user_passes_test(es_admin)
def gestionar_votaciones(request):
    votacion_id = request.POST.get("votacion_id") or request.GET.get("votacion_id")
    if votacion_id:
        votacion = Votacion.objects.get(IdVotacion=votacion_id)
        censos = Censo.objects.all()
        form1 = NuevaCandidaturaForm()
        mesa_ya_sorteada = MesaElectoral.objects.filter(IdVotacion=votacion).exists()
        candidaturas_votacion = Candidatura.objects.filter(candidatosnombrados__votacion=votacion)
        integrantes_asociados={}
        for candidatura in candidaturas_votacion:
            integrantes = IntegrantesCandidatura.objects.filter(candidatura=candidatura).select_related('usuario')
            integrantes_asociados[candidatura.IdCandidatura] = [integra.usuario.username for integra in integrantes]
        return render(request, "ManejarVotacion.html", {'form1': form1,'censos': censos,'votacion': votacion,'mesa_ya_sorteada': mesa_ya_sorteada, 'candidaturas':candidaturas_votacion, 'integrantes':integrantes_asociados})
    
    messages.error("No has seleccionado ninguna votacion")
    return redirect('administracion')

